import pandas as pd 
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import datetime
import nltk
import xlrd
import win32com.client as win32
from nltk import tokenize
import requests
from bs4 import BeautifulSoup
from lxml import etree
from lxml import html
import re
import time
import ast
from openpyxl.styles import PatternFill
import numpy as np
import sys
import unicodedata
from rich.console import Console
from rich.text import Text
# from sql import *
import streamlit as st
from io import StringIO, BytesIO
from PIL import Image
from xls2xlsx import XLS2XLSX
import os
import pythoncom
import tempfile
# from converter import convert_to_xlsx

# Set page configuration
st.set_page_config(layout="wide")

m = st.markdown("""
<style>
div.stButton > button:first-child {
    background-color: rgb(255,140,0);
    color: white;
}
div.stButton:hover > button:first-child {
border-color: white;
}
div.stButton:click > button:first-child {
border-color: white;
color: white;
}
a:link {
color: white;
text-decoration: none;
}
button.st-emotion-cache-1lp7pgu:hover {
    color: rgb(255,140,0);
}
button.st-emotion-cache-19rxjzo:hover {
    color: rgb(255,140,0);
    border-color: rgb(255,140,0);
}
button.st-emotion-cache-19rxjzo:active {
    color: rgb(255,140,0);
    border-color: rgb(255,140,0);
    text-clor: white;
    background-color: rgb(14, 17, 23);
    
}
button.st-emotion-cache-19rxjzo:focus:not(:active) {
    color: rgb(255,140,0);
    border-color: rgb(255,140,0);
    background-color: rgb(14, 17, 23);
    text-color: rgb(255,140,0);
}
button.st-emotion-cache-19rxjzo:clicked {
    color: rgb(255,140,0);
    border-color: rgb(255,140,0);
    background-color: rgb(255,140,0);
    text-clor: rgb(255,140,0);
}
.st-emotion-cache-p5msec:hover {
    color: rgb(255,140,0);
}
.st-emotion-cache-p5msec:hover svg {
fill: rgb(255,140,0);
}
</style>
                """, unsafe_allow_html=True)

image1 = Image.open("images/SimboloNEO_white-border_transp.png", 'r')

width1 = 100
width2 = 180
height = 80
image1_resized = image1.resize((int(image1.width * (height / image1.height)), height))
# image2_resized = image2.resize((int(image2.width * (height / image2.height)), height))

def send_email(link):
    '''
    Function to send an email to the support team
    '''
    js = f"window.open('{link}', '_blank')"
    html = f"<script>{js}</script>"
    st.markdown(html, unsafe_allow_html=True)


image = 'streamlit_app\images\TelefoneIcon.png'

st.sidebar.markdown('<h1 style="text-align: center; margin-bottom: 1px; padding-bottom: 5px">Contato</h1><hr style="margin-top: 10px" />', unsafe_allow_html=True)
st.sidebar.markdown('<div><h2 style="text-align: left;">📞 +55 (48) 3239-2041</h2></div>', unsafe_allow_html=True)
st.sidebar.markdown('<div><h2 style="text-align: left;"><a href="mailto:faleconosco@neo.certi.org.br">📩 faleconosco@neo.certi.org.br</a></h2></div>', unsafe_allow_html=True)

col1, col2 = st.columns([15, 3], gap="medium")

with col2:
    st.image([image1_resized])

with col1:
    st.title('Verificação de Certificados')

def convert_to_xlsx(file_path):
    excel = win32.gencache.EnsureDispatch('Excel.Application', pythoncom.CoInitialize())
    try:
        wb = excel.Workbooks.Open(file_path)
        output_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.SaveAs(output_path, FileFormat=51)
        wb.Close()
        excel.Quit()
        return output_path
    except Exception as e:
        print("Could not convert file:", file_path)
        print("Error:", e)
        return None

file_placeholder = st.empty()

file = file_placeholder.file_uploader('Escolha o arquivo que deseja verificar', type=('xls', 'xlsx'))

if file is not None:
    file_extension = os.path.splitext(file.name)[1]
    if file_extension != ".xlsx":
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False)
        temp_file.write(file.getvalue())
        temp_file.close()

        # Convert to xlsx
        converted_file_path = convert_to_xlsx(temp_file.name)
        if converted_file_path is not None:
            st.success("File conversion complete. Download the converted file below.")
            workbook = openpyxl.load_workbook(converted_file_path, data_only=True)
        else:
            st.error("Failed to convert the file.")

        # Clean up temporary file
        os.unlink(temp_file.name)
    else:
        workbook = openpyxl.load_workbook(file, data_only=True)
        st.warning("The uploaded file is already in .xlsx format.")

with st.expander('Instruções'):
    st.markdown('<h3>Como utilizar a ferramenta:</h3>', unsafe_allow_html=True)
    st.markdown('<p>1. Clique no botão acima para fazer o upload do arquivo que deseja verificar.</p>', unsafe_allow_html=True)
    st.markdown('<p>2. Aguarde enquanto o arquivo é carregado.</p>', unsafe_allow_html=True)
    st.markdown('<p>3. Após o carregamento, a ferramenta irá verificar se o arquivo está correto e, em caso positivo, irá gerar um relatório com os resultados.</p>', unsafe_allow_html=True)
    st.markdown('<p>4. Caso o arquivo não esteja correto, a ferramenta irá informar o erro encontrado.</p>', unsafe_allow_html=True)
    st.markdown('<p>5. Em caso de dúvidas, entre em contato com o suporte.</p>', unsafe_allow_html=True)
st.markdown('<hr />', unsafe_allow_html=True)

while file is None:
    time.sleep(1)

if file is not None:
    file_name = file.name
    st.markdown(f'<div style="text-align: center; font-size: 36px">{file_name}</div>', unsafe_allow_html=True)

# try:
#     workbook = openpyxl.load_workbook(file, data_only=True)
# except Exception as e:
#     st.error(f'Erro ao carregar o arquivo: Arquivo corrompido ou formato inválido.')
#     st.stop()

#Loading the worksheet (Excel file that we gonna work with)
sheet = workbook.active

# Get the max row count of the worksheet
max_row = sheet.max_row

# Get the max column count of the worksheet
max_column = sheet.max_column

def get_last_row(sheet):
    '''
    Function to get the last line with content on the Excel file
    '''
    for i in range(max_row, 0, -1):
        row_values = [cell.value for cell in sheet[i]]

        if any(row_values):
            return i
        
    return None

#Activating the function and storing the result in a variable (last_row)
last_row = get_last_row(sheet)

def get_last_column(sheet):
    '''
    Function to get the last column with content on the Excel file
    '''	
    for i in range(max_column, 0, -1):
        column_values = [cell.value for cell in sheet[i]]

        if any(column_values):
            return i
        
    return None

#Activating the function and storing the result in a variable (last_column)
last_column = get_last_column(sheet)

#Starting to create the main function
def main():
    ''' Function to analyze all data from tables

    Inner Functions:
    - scrapper()
    - convert_to_float()
    - process_string()
    - create_df_capa()
    - process_cmc_information()
    - convert_to_meters()
    - convert_to_mm()
    - convert_to_µm()
    - find_excel_row_by_value()

    Returns:
    - tables
    '''
    #Declaring the global variables (to use in other functions)
    global start_row
    global tables
    global df_padroes

    def scrapper():
        '''
        Function to scrape all data from RBC website
        '''
        #URL that we gonna scrape
        url = 'http://www.inmetro.gov.br/laboratorios/rbc/detalhe_laboratorio.asp?num_certificado=34&situacao=AT&area=DIMENSIONAL'

        #Requesting the URL
        response = requests.get(url)

        #Getting the HTML content
        html_content = response.content

        #Using BeatifulSoup to parse the HTML
        soup = BeautifulSoup(html_content, 'html.parser')

        #Here, we use the lxml to parse the HTML (it's a better parser than the default one from BeautifulSoup)
        html_tree = html.fromstring(str(soup))

        #Using XPATH to find the table that we want
        table_rows = html_tree.xpath('//table[4]/tr')

        #Creating a list to store the data
        rows_data = []

        #Iterating through the rows of the table and getting the data (text format)
        for row in table_rows:
            cells = row.xpath('.//td|.//th')
            row_data = [cell.text_content().strip() for cell in cells]
            rows_data.append(row_data)

        #Creating a DataFrame to store the data (df)
        df = pd.DataFrame(rows_data, columns = None, index = None) 

        #Dropping rows and columns with all NaN values
        df = df.dropna(axis=1, how='all')
        df = df.dropna(axis=0, how='all')

        #Changing values from df to match them with the information we gonna get at the Excel file
        df = df.replace('Medição de', 'Medir', regex=True)
        df = df.replace('Medição por', 'Medir por', regex=True)
        df = df.replace('para Medir', 'de Medir', regex=True)
        df = df.replace('Medição', 'Medir', regex=True)
        df.columns = ['Descrição do serviço', 'Parâmetro, Faixa e Método', 'Capacidade de Medição e Calibração (CMC)']

        #Treating data
        for i, row in df.iterrows():
            if pd.isna(df.at[i, 'Descrição do serviço']) or row['Descrição do serviço'] == '':
                df.at[i, 'Descrição do serviço'] = df.at[i-1, 'Descrição do serviço']

        #Returning df (that's the output of my scrapper() function)
        return df

    #Here, we gonna list some function to use afterwards

    #Function to convert the values to float (we can apply any time we need it)
    def convert_to_float(value):
        try:
            return float(value.split()[0].replace(',', '.'))
        except:
            return value

    #Function to process the string and get the interval
    def process_string(string):
        result = re.findall(r'\d+', string)

        if result:
            if len(result) != 1:
                return float(result[0]), float(result[1])
            else:
                return 0.0, float(result[0])
        else:
            return None

    #Extracting data from Excel file

    #Getting the data from the first part of the Excel file (the cover of the document)
    centro_found = False
    padroes_found = False
    capa_data = []

    #Iterating through the rows of the Excel file
    for i in range (1, last_row + 1):
        row_values = []
        #Iterating through the columns of the Excel file
        for j in range (1, last_column + 1):
            #Getting the value of the cell and storing on cell_obj (the value of a cell is the content of it)
            cell_obj = sheet.cell(row = i, column = j)

            #Creating a logic to find the start and finish of the cover of the document
            #The values of the cells are going to be used as the , knowing that all certificates have the same structure
            if str(cell_obj.value).startswith('CENTRO'):
                centro_found = True
                continue
            if str(cell_obj.value).startswith('Padrões utilizados'):
                padroes_found = True
                continue

            #Appeding cell values into a list (row_values)
            if centro_found and not padroes_found:
                row_values.append(cell_obj.value)
                if len(row_values) == 9:
                    capa_data.append(row_values)

        if padroes_found:
            break

    def create_df_capa():
        '''
        Function to create df_capa, using the capa_data list
        '''
        df_capa = pd.DataFrame(capa_data)
        df_capa = df_capa.dropna(axis=1, how='all')
        df_capa = df_capa.dropna(axis=0, how='all')
        df_capa = df_capa[~df_capa.apply(lambda row: 'Ocultar' in row.values, axis=1)]
        df_capa = df_capa.drop_duplicates().reset_index(drop=True)

        return df_capa

    df_capa = create_df_capa()

    #Extracting the machines listed on the Excel file

    padroes_found = False
    procedimento_found = False
    padroes_data = []

    for i in range(1, last_row + 1):
        row_values = []
        for j in range(1, last_column + 1):
            cell_obj = sheet.cell(row = i, column = j)

            if str(cell_obj.value) == 'Padrões utilizados':
                padroes_found = True
                continue

            if str(cell_obj.value).startswith('Procedimento'):
                procedimento_found = True
                continue

            if padroes_found and not procedimento_found:
                row_values.append(cell_obj.value)
                if len(row_values) == 9:
                    padroes_data.append(row_values)

        if procedimento_found:
            break

    #Creating a DataFrame to store the machine data (df_padroes) and treating data
    df_padroes = pd.DataFrame(padroes_data)
    df_padroes = df_padroes.dropna(axis=1, how='all')

    descricao_column = df_padroes[df_padroes.eq('Descrição').any(axis = 1)].stack().index[1][1]

    start_row = df_padroes[df_padroes[descricao_column] == 'Descrição'].index[0]
    value_below_descricao = df_padroes.loc[start_row + 1][descricao_column]

    result_values = [value_below_descricao]
    current_row = start_row + 1

    while current_row < len(df_padroes) and df_padroes.loc[current_row, descricao_column] is not None:
        result_values.append(df_padroes.loc[current_row, descricao_column])
        current_row += 1

    machines_df = pd.DataFrame(result_values, columns=[descricao_column])

    machines_df = machines_df.dropna(how='all', inplace=False)
    machines_df = machines_df.drop_duplicates().reset_index(drop=True)

    machines_df = machines_df[machines_df[descricao_column] != '#N/A']

    #Naming columns and finalizing the creation machines_df
    machines_df.columns = ['Descrição do serviço']

    #Extracting the data inside tables on the Excel file
    resultados_found = False
    observacoes_found = False
    data = []

    for i in range(1, last_row + 1):
        row_values = []

        for j in range(1, last_column + 1):
            cell_obj = sheet.cell(row = i, column = j)

            if str(cell_obj.value) == 'Resultados':
                resultados_found = True
                continue
        
            if str(cell_obj.value) == 'Observações':
                observacoes_found = True
                continue

            if resultados_found and not observacoes_found:
                row_values.append(cell_obj.value)
                if len(row_values) == 9:
                    data.append(row_values)

        if observacoes_found:
            break

    #Creating a DataFrame to store the table data (df_dados) and treating data
    df_dados = pd.DataFrame(data)
    df_dados = df_dados.dropna(axis=1, how='all')
    df_dados = df_dados.dropna(axis=0, how='all')
    df_dados = df_dados[~df_dados.apply(lambda row: 'Ocultar' in row.values, axis=1)]
    df_dados = df_dados.drop_duplicates().reset_index(drop=True)

    #Separating the different tables inside the df_dados (the number of tables change according to the document)

    new_table_indices = df_dados[df_dados.apply(lambda row: any(cell and str(cell).startswith('Valor') for cell in row.values), axis=1)].index

    tables = []

    for i in range(len(new_table_indices)):
        start_idx = new_table_indices[i]
        end_idx = new_table_indices[i+1] if i+1 < len(new_table_indices) else len(df_dados)
        table = df_dados.iloc[start_idx:end_idx, :].reset_index(drop=True)
        tables.append(table)

    #The tables list is going to store all the tables we have inside the df_dados

    #Calling the scrapper() function and storing the result in a variable (df_web)
    df_web = scrapper()

    #Merging the machines_df with df_web (same result as a VLOOKUP in Excel) at df_merge

    df_merge = pd.merge(machines_df, df_web, on='Descrição do serviço', how='left')
    df_merge = df_merge.drop_duplicates().reset_index(drop=True)

    #Coping previous DataFrames to use them later
    df_capa_merge = df_capa.copy()
    df_web_merge = df_web.copy()

    #Changing the values of the columns to lowercase, to avoid problems merging the DataFrames
    df_capa_merge[df_capa_merge.columns[1]] = df_capa_merge[df_capa_merge.columns[1]].str.lower()
    df_web_merge[df_web_merge.columns[0]] = df_web_merge[df_web_merge.columns[0]].str.lower()

    #Merging the df_capa_merge with df_web_merge (same result as a VLOOKUP in Excel) at df_service
    df_service = pd.merge(df_capa_merge, df_web_merge, left_on=df_capa_merge.iloc[:, 1], right_on=df_web_merge.iloc[:, 0], how='inner')
    df_service = df_service.drop_duplicates().reset_index(drop=True)

    #Locating the first column of tables and converting the values to float
    first_column = tables[0].iloc[:, 0]
    for i, table in enumerate(tables):
        tables[i].iloc[:, 0] = table.iloc[:, 0].apply(convert_to_float)

    first_column.columns = ['Resultados']

    last_value = first_column.iloc[-1]

    #Creating a logic to separate works done in the field and in the laboratory
    #The base of the logic is that, at the cover, a cell with the value 'LOCAL DA CALIBRAÇÃO' means that the work was done in the field

    #Copping the df_web to use it
    df_web_split = df_web.copy()

    #Finding the index of the cell with the value 'LOCAL DA CALIBRAÇÃO'
    indices = df_web_split[df_web_split['Descrição do serviço'] == 'INSTRUMENTOS E GABARITOS DE MEDIÇÃO DE ÂNGULO'].index[1]

    df_web_lab = df_web_split.iloc[:indices - 2]
    df_web_field = df_web_split.iloc[indices - 2:]

    #Finding if the work was done in the field or in the laboratory

    if 'LOCAL DA CALIBRAÇÃO' in df_capa.iloc[:, 0].astype(str).values:
        working_df = df_web_field
    else:
        working_df = df_web_lab

    #Merging the df_cap_merge and working_df, to get the values from RBC based on the work that is listed on the Excel file

    working_df[working_df.columns[0]] = working_df[working_df.columns[0]].str.lower()

    df_merge_service = pd.merge(df_capa_merge, working_df, left_on=df_capa_merge.iloc[:, 1], right_on=working_df.iloc[:, 0], how='inner')

    #Treating data
    df_merge_service = df_merge_service.drop_duplicates().reset_index(drop=True)
    df_merge_service.iloc[:, -1] = df_merge_service.iloc[:, -1].str.replace('*', '')
    df_merge_service = df_merge_service.dropna(axis=0, how='all')
    df_merge_service = df_merge_service.dropna(axis=1, how='all')

    #Extracting the numerical range from each line of tables
    try:
        df_merge_service['Intervalo'] = df_merge_service.iloc[:, 4].apply(process_string)
        df_merge_service = df_merge_service.dropna(axis=0, how='any')
        df_merge_service = df_merge_service.dropna(axis=1, how='any')  

    except Exception as e:
        print(f'{e}: Não foi possível encontrar a máquina utilizada para o serviço')

    #Creating a logic to get each value of the first column, and with that discover the correct numerical range

    def process_cmc_information(cmc_value):
        '''
        Function to process the CMC information, based on the string format
        '''
        cmc_value = str(cmc_value).strip()
        
        # Check if it's a distance (type 1)
        distance_match = re.match(r'([\d.,]+)\s*([µm]+)', cmc_value)
        if distance_match:
            value = float(distance_match.group(1).replace(',', '.'))
            unit = distance_match.group(2)
            return value, unit

        # Check if it's an equation (type 2)
        equation_match = re.match(r'\[([\s\S]+)\]', cmc_value)
        if equation_match:
            return equation_match.group(1)

        # Check if it's an angle (type 3)
        angle_match = re.match(r'\s*(\d+)\s*\'\'\s*', cmc_value)
        if angle_match:
            return float(angle_match.group(1))

        # Check if it's a percentage (type 4)
        percentage_match = re.match(r'([\d.,]+)%', cmc_value)
        if percentage_match:
            return float(percentage_match.group(1).replace(',', '.')) / 100

        # Default case: return the original value
        return cmc_value

    #Applying the function to the column 'Capacidade de Medição e Calibração (CMC)'
    df_merge_service['Capacidade de Medição e Calibração (CMC)'] = df_merge_service['Capacidade de Medição e Calibração (CMC)'].apply(process_cmc_information)

    def get_error_and_uncertainty(valor, intervalos):
        for intervalo in intervalos:
            if intervalo[0] <= valor <= intervalo[1]:
                return intervalo[0]
        return None

    single_row = df_merge_service.iloc[0]

    # Iterating through each table in the 'tables' list
    for i, row in df_merge_service.iterrows():
        # Iterating through each table in the 'tables' list
        for j, table in enumerate(tables):
            # Getting the numerical range for the current row and table
            intervalo = tuple(row['Intervalo'])
            
            # Create a new column with True if the value is within the range, False otherwise
            table[f'Within_Range_{i + 1}'] = pd.to_numeric(table.iloc[:, 0], errors='coerce').between(*intervalo)

    #Co-relating df_merge_service with tables

    for i, table in enumerate(tables):
        # Create a new column to store the selected values
        table['Selected_Value'] = ""

        # Iterate through the rows of the table
        for index, row in table.iterrows():
            selected_value = ""

            # Iterate through the columns and find the first True value in boolean columns
            for col in table.columns:
                if isinstance(col, str) and col.startswith('Within_Range') and row[col]:
                    # Extract the range index from the column name
                    range_index = int(col.split('_')[-1]) - 1

                    # Check if the range_index is within the valid range of rows in df
                    if 0 <= range_index < len(df_merge_service):
                        # Get the value from the 6th column of df based on the range_index
                        selected_value = df_merge_service.iloc[range_index, -2]  # Assuming the 6th column index is 5
                        break  # Exit the loop once a match is found

            # Update the 'Selected_Value' column with the corresponding value from df
            table.at[index, 'Selected_Value'] = selected_value

    #Separating the values of error and uncertainty on a new column

    for i, table in enumerate(tables):
        u_column_index = table.columns[table.iloc[0].astype(str).str.startswith('U')].tolist()
        if u_column_index:
            u_column_index = u_column_index[0]
            break
        
    has_meters = '[m]' in table.iloc[:, u_column_index].values
    has_mm = '[mm]' in table.iloc[:, u_column_index].values
    has_µm = '[µm]' in table.iloc[:, u_column_index].values

    for i in range(len(tables)):
        table = tables[i]

        if 'Selected_Value' in table.columns:
            table[['CMC_Value', 'CMC_Unit']] = table['Selected_Value'].apply(pd.Series)

            table = table.drop('Selected_Value', axis=1)

            tables[i] = table
        else:
            pass

    #Working with the units of the values

    def convert_to_meters(row):
        '''
        Function to convert the values to meters
        '''
        value, unit = row['CMC_Value'], row['CMC_Unit']
        if has_meters and unit == 'mm':
            return value / 1000  
        elif has_meters and unit == 'µm':
            return value / 1000000  
        else:
            return value

    def convert_to_mm(row):
        '''
        Function to convert the values to millimeters
        '''
        value, unit = row['CMC_Value'], row['CMC_Unit']
        if has_mm and unit == 'm':
            return value * 1000  
        elif has_mm and unit == 'µm':
            return value / 1000  
        else:
            return value

    def convert_to_µm(row):
        '''
        Function to convert the values to micrometers
        '''
        value, unit = row['CMC_Value'], row['CMC_Unit']
        if has_µm and unit == 'm':
            return value * 1000000  
        elif has_µm and unit == 'mm':
            return value / 1000  
        else:
            return value
        
    #Converting the values to the respective correct unit
    for i in range(len(tables)):
        table = tables[i]
        if 'CMC_Value' in table.columns and 'CMC_Unit' in table.columns and has_meters:
            table['CMC_Value'] = table.apply(convert_to_meters, axis=1)
            tables[i] = table
        elif 'CMC_Value' in table.columns and 'CMC_Unit' in table.columns and has_mm:
            table['CMC_Value'] = table.apply(convert_to_mm, axis=1)
            tables[i] = table
        elif 'CMC_Value' in table.columns and 'CMC_Unit' in table.columns and has_µm:
            table['CMC_Value'] = table.apply(convert_to_µm, axis=1)
            tables[i] = table
        else:
            pass

    for i, table in enumerate(tables):
        tables[i].iloc[:, u_column_index] = table.iloc[:, u_column_index].apply(convert_to_float)

    #Creating new columns to store results from later checks
    for i in range(len(tables)):
        table = tables[i]
        table['CMC_Value'] = table['CMC_Value'].replace('', None)
        table['CMC_Verification'] = None
        table['Range_Verification'] = None
        table['Correction_Verification'] = None

    #Searching for the resolution row on the df_capa DataFrame
    search_condition = (df_capa[0] == 'RESOLUÇÃO')
    result = df_capa.loc[search_condition]

    #Chekicking if the result is not empty, and getting the resolution value (as an string)
    if not result.empty:
        resolucao_value = result.iloc[0, 1]

        #Converting the resolution value to float
        resolucao_value = resolucao_value.split()[0]
        resolucao_value = resolucao_value.replace(',', '.')
        #Multiplying the resolution value by 3
        resolucao_value = float(resolucao_value) * 3
        #Creating the negative value of the resolution
        resolucao_negative_value = -resolucao_value

        #Creating a logic to check if the values are within the resolution range
        for i, table in enumerate(tables):
            for index, row in table.iterrows():
                correction_column_value = row.iloc[2]

                # Check if the value is not None and not NaN
                if pd.notna(correction_column_value) and correction_column_value != 'Correção':
                    
                    if type(correction_column_value) == str:
                        correction_column_value = pd.to_numeric(
                            correction_column_value.replace(',', '.'),
                            errors='coerce'
                            )
                    else:
                        pass
                    
                    if pd.notna(correction_column_value): # Check if the value is not None and not NaN
                        # Check if the value is within the resolution range
                        if correction_column_value < resolucao_negative_value or correction_column_value > resolucao_value:
                            table.at[index, 'Correction_Verification'] = True
                
                    if pd.notna(correction_column_value):  # Check again after conversion
                        
                        if correction_column_value < resolucao_negative_value or correction_column_value > resolucao_value:

                            table.at[index, 'Correction_Verification'] = True
                else:
                    pass
    else:
        pass

    for i, table in enumerate(tables):
        for index, row in table.iterrows():
            u_column_value = row[u_column_index]
            cmc_value = row['CMC_Value']

            if pd.notna(cmc_value) and pd.notna(u_column_value):
                if u_column_value >= cmc_value:
                    table.at[index, 'CMC_Verification'] = False
                elif u_column_value  < cmc_value:
                    table.at[index, 'CMC_Verification'] = True
            elif pd.isna(cmc_value) and type(u_column_value) != str and pd.notna(u_column_value):
                table.at[index, 'Range_Verification'] = True


    def find_excel_row_by_value_range(sheet, target_value):
        '''
        Function to find the row in the Excel file that corresponds to the target value
        '''
        max_row = sheet.max_row

        # Iterating through rows in the Excel sheet
        for i in range(1, max_row + 1):
            cell_obj = sheet.cell(row=i, column=1)  

            # Checking if the cell contains the target value
            if cell_obj.value == target_value:
                return i  # Return the row number if found

        return None  # Return None if not found
    
    def find_excel_row_by_value_cmc(sheet, target_value):
        '''
        Function to find the row in the Excel file that corresponds to the target value
        '''
        max_row = sheet.max_row

        # Iterating through rows in the Excel sheet
        for i in range(1, max_row + 1):
            cell_obj = sheet.cell(row=i, column=4)  

            # Checking if the cell contains the target value
            if cell_obj.value == target_value:
                return i  # Return the row number if found

        return None  # Return None if not found

    # Iterating through each DataFrame in the list
    list=[]

    cell_coordinate=[]

    cell_coordinate_cmc=[]
    cell_coordinate_error=[]

    target_value=[]
    target_value_cmc_list=[]
    target_value_range_list=[]

    range_error=[]
    cmc_error=[]


    for i, table in enumerate(tables):

        for index, row in table.iterrows():
            # Assuming the first column in the table corresponds to the second column in Excel

            if row['Range_Verification'] == True or row['CMC_Verification'] == True:
                target_value_range = row[0]
                target_value_cmc = row[3]
                target_value_str = str(target_value_range).replace('.', ',')
            # Finding the row in the Excel file that corresponds to the target value

                if target_value_range is not None or target_value_cmc is not None:

                    if row['Range_Verification'] == True:
                        excel_row = find_excel_row_by_value_range(sheet, target_value_str)
                        list.append(f'<span style="font-size:20px;">Erro de range na linha {excel_row} - {target_value_str} - valor fora do range indicado pela RBC</span><br>')
                        cell_coordinate_error.append(excel_row)
                        target_value.append(target_value_str)
                        target_value_range_list.append(target_value_str)
                        range_error.append('Erro de range')

                    elif row['CMC_Verification'] == True:
                        excel_row = find_excel_row_by_value_cmc(sheet, target_value_cmc)
                        list.append(f'<span style="font-size:20px;">Erro de CMC na linha {excel_row} - {target_value_cmc} - valor fora do CMC indicado pela RBC</span><br>')
                        cell_coordinate_cmc.append(excel_row)
                        target_value.append(target_value_cmc)
                        target_value_cmc_list.append(target_value_cmc)
                        cmc_error.append('Erro de CMC')


                else:
                    pass
                    
            else:
                pass

    if list:
        col1, col2, col3 = st.columns([1, 1.2, 1])


    if cell_coordinate_error:
        data = {'Linha': cell_coordinate_error, 'Valor': target_value_range_list, 'Erro': range_error}
        df_range = pd.DataFrame(data)

    if cell_coordinate_cmc:
        data = {'Linha': cell_coordinate_cmc, 'Valor': target_value_cmc_list, 'Erro': cmc_error}
        df_cmc = pd.DataFrame(data)

    final_df = pd.concat([df_range, df_cmc], axis=0)
    final_df = final_df.sort_values(by='Linha', ascending=True)

    with col2:
        st.markdown('<h2 style="text-align: center;">Erros na tabela de medição:</h2></br>', unsafe_allow_html=True)
        st.markdown(
    '<div style="display: flex; justify-content: center; margin-bottom: 20px;">' + 
    final_df.style.hide(axis="index").to_html() +
    '</div><hr />', 
    unsafe_allow_html=True
)
    return tables

error_list = []



def verify_pattern_origin():
    '''
    Function to verify the origin of a service.

    The origin is listed at the 'Padrões utilizados' section of the document.

    If the value is CERTI, the function returns a message saying that the certificate is from CERTI.

    If the value is LMD, the function returns a message saying that the certificate is not from CERTI (that's a problem).
    '''
    standards_df = df_padroes.copy()
    standards_df = standards_df.dropna(axis=1, how='all').reset_index(drop=True)
    standards_df = standards_df.dropna(axis=0, how='all').reset_index(drop=True)

    origin_column = standards_df.iloc[:, -2]

    if origin_column[1] == 'LMD':
        error_list.append('<span style="font-size:16px;">Erro: O certificado não é do padrão CERTI</span><br>')
    else:
        pass

def verify_pattern_alignment():
    '''
    Function to verify the text alignment of the 'Padrões utilizados' section of the document.
    '''
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value is not None:
                if cell.value == 'Padrões utilizados':
                    start_row = cell.row + 1  # Start from the next row
                elif cell.value.startswith('Procedimento de'):
                    end_row = cell.row - 1  # End at the previous row
                    break

    # Check alignment and content for each cell in the range
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                # Get alignment properties
                alignment = cell.alignment
                horizontal_alignment = alignment.horizontal
                vertical_alignment = alignment.vertical
                if horizontal_alignment == 'left' or horizontal_alignment == None:
                    if vertical_alignment == 'center' or vertical_alignment == None:
                        pass
                    else:
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span><span> Alinhanento vertical incorreto</span><br>')
                        
                else:
                    error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Alinhamento horizontal incorreto</span><br>')

verify_pattern_font_coordinates = []

verify_pattern_font_fonts = []

def verify_pattern_font():
    '''
    Function to verify the font of the 'Padrões utilizados' section of the document.
    '''
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Padrões utilizados':
                start_row = cell.row + 2  # Start from the next row
            elif cell.value == 'Procedimento de calibração':
                end_row = cell.row - 1  # End at the previous row
                break

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                # Get alignment properties
                font_name = cell.font.name
                font_size = cell.font.sz
                if font_name == 'Nunito Sans':
                    if cell.row == start_row and font_size == 10:
                        pass
                    elif cell.row == start_row and font_size == 9:
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Fonte {font_size} incorreta.</span><br>')
                        verify_pattern_font_coordinates.append(cell.coordinate)
                        verify_pattern_font_fonts.append(font_size)
                    elif cell.row != start_row and font_size == 9:
                        pass
                    else:
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Fonte e tamanho atuais: {font_name, font_size} Correto: Nunito Sans, 9.0</span><br>')
                        verify_pattern_font_coordinates.append(cell.coordinate)
                        verify_pattern_font_fonts.append(font_size)
                else:
                    error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> fonte e tamanho atuais: {font_name, font_size} Correto: Nunito Sans, 9.0</span><br>')
                    verify_pattern_font_coordinates.append(cell.coordinate)
                    verify_pattern_font_fonts.append(font_size)

verify_procedure_text_font_coordinates = []

verify_procedure_text_font_font_size = []

verify_procedure_text_font_font_name = []

def verify_procedure_text_font():
    '''
    Function to verify the font of the text in the 'Procedimento de calibração' section of the document.

    All the cells that contain text at this section should have the font 'Nunito Sans' and the size 10.
    '''
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value is not None:
                if cell.value.startswith('Procedimento de'):
                    start_row = cell.row + 1  # Start from the next row
                elif cell.value == 'Resultados':
                    end_row = cell.row - 1  # End at the previous row
                    break

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                font_name = cell.font.name
                font_size = cell.font.sz
                if font_name == 'Nunito Sans' and font_size == 10:
                    pass
                else:
                    error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span style="font-size:16px;"> Fonte {font_name, font_size} incorreta. Correto: Nunito Sans, 10.0</span><br>')
                    verify_procedure_text_font_coordinates.append(cell.coordinate)
                    verify_procedure_text_font_font_size.append(font_size)
                    verify_procedure_text_font_font_name.append(font_name)

verify_titles_coordinates = []

verify_titles_font_names = []

verify_titles_font_sizes = []

verify_titles_font_bold = []
def verify_titles():
    '''
    Function to verify the font of all titles in the document.

    The titles are:
    - Padrões utilizados
    - Procedimento de calibração
    - Resultados
    - Observações

    '''
    titles = []
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Padrões utilizados':
                titles.append(cell.coordinate)
            elif cell.value == 'Procedimento de calibração':
                titles.append(cell.coordinate)
            elif cell.value == 'Resultados':
                titles.append(cell.coordinate)
            elif cell.value == 'Observações':
                titles.append(cell.coordinate)
            else:
                pass
    for item in titles:
        font_name = sheet[item].font.name
        font_size = sheet[item].font.sz
        font_bold = sheet[item].font.b
        if font_name == 'Nunito Sans' and font_size == 11 and font_bold == True:
            pass
        elif font_name != 'Nunito Sans':
            error_list.append(f'<span style="font-size:26px;">{item}</span>:<span> Fonte icorreta. Correta: Nunito Sans</span><br>')
        elif font_size != 11:
            error_list.append(f'<span style="font-size:26px;">{item}</span>:<span> Tamanho da fonte incorreto. Correto: 11></span><br>')
        elif font_bold != True:
            error_list.append(f'<span style="font-size:26px;">{item}</span>:<span> Erro no negrito. Estado atual:{font_bold}, espera-se True</span><br>')
        else:
            pass

def verify_observations_text():
    '''
    Function to verify the font of the text in the 'Observações' section of the document.

    All the cells that contain text at this section should have the font 'Nunito Sans' and the size 9.
    '''
    start_row = None
    end_row = get_last_row(sheet)

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Observações':
                start_row = cell.row + 1
                pass
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                # Get alignment properties
                font_name = cell.font.name
                font_size = cell.font.sz
                if font_name == 'Nunito Sans' and font_size == 9:
                    pass
                else:
                    error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Observações com formato {font_name,font_size} incorreto. Correto: Nunito Sans, 9.0</span><br>')

def verify_executer_font():
    '''
    Function to verify the font of the text in the 'Executor' section of the document.

    All the cells that contain text at this section should have the font 'Nunito Sans', the size 10 and bold.
    '''
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value and str(cell.value).startswith('Executor'):
                start_row = cell.row
            elif cell.value == 'Padrões utilizados':
                end_row = cell.row - 1
                break
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not 
            if cell.value is not None:
                if not str(cell.value).startswith('Executor'):
                    font_name = cell.font.name
                    font_size = cell.font.sz
                    font_bold = cell.font.b
                    if font_name == 'Nunito Sans' and font_size == 10 and font_bold == True:
                        pass
                    elif font_name != 'Nunito Sans':
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Fonte {font_name} incorreta. Correto: Nunito Sans</span><br>')
                    elif font_size != 10:
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Tamanho da fonte {font_size} incorreto. Correto: 10.0</span><br>')
                    elif font_bold != True:
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Estado de negrito {font_bold} incorreto, espera-se True</span><br>')
                    else:
                        pass

def verify_table_observation_font():
    '''
    Function to verify the font of the text in the 'Observações' section of the document.

    All the cells that contain text at this section should have the font 'Nunito Sans' and the size 8.
    '''
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Resultados':
                start_row = cell.row + 1
            elif cell.value == 'Observações':
                end_row = cell.row - 1
                break
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                if str(cell.value).startswith('Obs'):
                    font_name = cell.font.name
                    font_size = cell.font.sz
                    if font_name == 'Nunito Sans' and font_size == 8:
                        pass
                    elif font_name != 'Nunito Sans':
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Fonte {font_name} incorreta. Correta: Nunito Sans</span><br>')
                    elif font_size != 8:
                        error_list.append(f'<span style="font-size:26px;">{cell.coordinate}</span>:<span> Tamanho da fonte {font_size} incorreto. Correto: 8.0</span><br>')
                    else:
                        pass

def verify_header():
    '''
    Function to verify the font of the header of the document.

    The header is the first 5 rows of the document.

    For each row of header, a different font and size is expected.
    '''
    start_row_height = sheet.row_dimensions[1].height
    second_row_height = sheet.row_dimensions[2].height
    third_row_height = sheet.row_dimensions[3].height
    fourth_row_height = sheet.row_dimensions[4].height
    fifth_row_height = sheet.row_dimensions[5].height

    start_row_name = sheet.cell(row=1, column=1).font.name
    start_row_size = sheet.cell(row=1, column=1).font.sz
    start_row_bold = sheet.cell(row=1, column=1).font.b

    second_row_name = sheet.cell(row=1 + 1, column=1).font.name
    second_row_size = sheet.cell(row=1 + 1, column=1).font.sz
    second_row_bold = sheet.cell(row=1 + 1, column=1).font.b

    third_row_name = sheet.cell(row=1 + 2, column=1).font.name
    third_row_size = sheet.cell(row=1 + 2, column=1).font.sz
    third_row_bold = sheet.cell(row=1 + 2, column=1).font.b

    fourth_row_name = sheet.cell(row=1 + 3, column=1).font.name
    fourth_row_size = sheet.cell(row=1 + 3, column=1).font.sz
    fourth_row_bold = sheet.cell(row=1 + 3, column=1).font.b

    fifth_row_name = sheet.cell(row=1 + 4, column=1).font.name
    fifth_row_size = sheet.cell(row=1 + 4, column=1).font.sz
    fifht_row_bold = sheet.cell(row=1 + 4, column=1).font.b

    if start_row_name == 'Nunito Sans' and start_row_size == 14 and start_row_height == 24.75 and start_row_bold != True:
        pass
    elif start_row_name != 'Nunito Sans':
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=start_row, column=1).coordinate}</span>,<span> fonte atual: {start_row_name} - fonte correta: Nunito Sans</span><br>')
    elif start_row_size != 14:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=start_row, column=1).coordinate}</span>,<span> tamanho atual: {start_row_size} - tamanho correto: 14.0</span><br>')
    elif start_row_height != 24.75:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=start_row, column=1).coordinate}</span>,<span> altura atual: {start_row_height} - altura correta: 24.75</span><br>')
    elif start_row_bold == True:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=start_row, column=1).coordinate}</span>,<span> negrito atual: {start_row_bold} - negrito correto: False</span><br>')
    else:
        pass

    if second_row_name == 'Nunito Sans' and second_row_size == 14 and second_row_height == 17.25 and second_row_bold != True:
        pass
    elif second_row_name != 'Nunito Sans':
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 1, column=1).coordinate}</span>,<span> fonte atual: {second_row_name} - fonte correta: Nunito Sans</span><br>')
    elif second_row_size != 14:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 1, column=1).coordinate}</span>,<span> tamanho atual: {second_row_size} - tamanho correto: 11.0</span><br>')
    elif second_row_height != 17.25:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 1, column=1).coordinate}</span>,<span> altura atual: {second_row_height} - altura correta: 18.0</span><br>')
    elif second_row_bold == True:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 1, column=1).coordinate}</span>,<span> negrito atual: {second_row_bold} - negrito correto: False</span><br>')
    else:
        pass

    if third_row_name == 'Nunito Sans' and third_row_size == 9 and third_row_height == 15 and third_row_bold != True:
        pass
    elif third_row_name != 'Nunito Sans':
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 2, column=1).coordinate}</span>,<span> fonte atual: {third_row_name} - fonte correta: Nunito Sans</span><br>')
    elif third_row_size != 9:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 2, column=1).coordinate}</span>,<span> tamanho atual: {third_row_size} - tamanho correto: 11.0</span><br>')
    elif third_row_height != 15:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 2, column=1).coordinate}</span>,<span> altura atual: {third_row_height} - altura correta: 18.0</span><br>')
    elif third_row_bold == True:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 2, column=1).coordinate}</span>,<span> negrito atual: {third_row_bold} - negrito correto: False</span><br>')
    else:
        pass

    if fourth_row_name == 'Nunito Sans' and fourth_row_size == 22 and fourth_row_height == 39 and fourth_row_bold == True:
        pass
    elif fourth_row_name != 'Nunito Sans':
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 3, column=1).coordinate}</span>,<span> fonte atual: {fourth_row_name} - fonte correta: Nunito Sans</span><br>')
    elif fourth_row_size != 22:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 3, column=1).coordinate}</span>,<span> tamanho atual: {fourth_row_size} - tamanho correto: 22.0</span><br>')
    elif fourth_row_height != 39:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 3, column=1).coordinate}</span>,<span> altura atual: {fourth_row_height} - altura correta: 39.0</span><br>')
    elif fourth_row_bold != True:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 3, column=1).coordinate}</span>,<span> negrito atual: {fourth_row_bold} - negrito correto: True</span><br>')
    else:
        pass

    if fifth_row_name == 'Nunito Sans' and fifth_row_size == 22 and fifth_row_height == 35.25 and fifht_row_bold == True:
        pass
    elif fifth_row_name != 'Nunito Sans':
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 4, column=1).coordinate}</span>,<span> fonte atual: {fifth_row_name} - fonte correta: Nunito Sans</span><br>')
    elif fifth_row_size != 22:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 4, column=1).coordinate}</span>,<span> tamanho atual: {fifth_row_size} - tamanho correto: 22.0</span><br>')
    elif fifth_row_height != 35.25:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 4, column=1).coordinate}</span>,<span> altura atual: {fifth_row_height} - altura correta: 35.25</span><br>')
    elif fifht_row_bold != True:
        error_list.append(f'<span style="font-size:26px;">{sheet.cell(row=1 + 4, column=1).coordinate}</span>,<span> negrito atual: {fifht_row_bold} - negrito correto: True</span><br>')
    else:
        pass

def verify_intern_procedure_code():
    '''
    Function to verify the internal procedure code of the document.

    The internal procedure code is the CMI number that is present in the 'Procedimento de calibração' section of the document.

    The CMI number is expected to be in the format 'CMI-XXX', where 'XXX' is a number.

    If the CMI number is 'CMI-000', the function returns an error message.

    If the CMI number is not found, the function returns an error message.
    '''
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Procedimento de calibração':
                start_row = cell.row + 1
            elif cell.value == 'Resultados':
                end_row = cell.row - 1
                break
    cell_value = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                cell_value.append(cell.value)
    words = cell_value[0].split()
    cmi_word = next((word for word in words if word.startswith('CMI')), None)
    if cmi_word is not None:
        cmi_word_parts = cmi_word.split('-')
        cmi_word_parts_number = cmi_word_parts[-1]
        if len(cmi_word_parts_number) > 3:
            cmi_word_parts_number = cmi_word_parts_number[:3]
        else:
            pass
        if cmi_word_parts_number == '000':
            error_list.append('<span style="font-size:16px;">Procedimento de calibração zerado</span><br>')
        else:
            pass
    else:
        error_list.append('<span style="font-size:16px;">Procedimento de calibração não encontrado (possível procedimento fantasma)</span><br>')

def verify_Veff():
    '''
    Function to verify the Veff value of the document.

    The Veff value is present in the 'Resultados' section of the document.

    If the Veff is 'Infinito', the k value need to be 2.
    '''
    for i, table in enumerate(tables):
        for index, row in table.iterrows():
            k_column_index = row[row == 'k'].index
            neff_column_index = row[row == 'neff'].index
            if not k_column_index.empty and not neff_column_index.empty:
                k_column_index = k_column_index.values
                neff_column_index = neff_column_index.values

                k_column_index = k_column_index[0]
                neff_column_index = neff_column_index[0]

                k_column = table.iloc[:, k_column_index]
                neff_column = table.iloc[:, neff_column_index]

                df = pd.DataFrame({'k': k_column, 'neff': neff_column})
                df = df.dropna(axis=0, how='any')
                df = df[df['k'] != 'k']
                df['Verification'] = None

                for index, row in df.iterrows():
                    if row['k'] == 2:
                        if row['neff'] == 'Infinito' or row['neff'] == 'infinito':
                            pass
                        else:
                            error_list.append(f'<span style="font-size:26px;">Linha {index} da tabela</span>:<span> νeff diferente de infinito para k = 2</span><br>')
                    else:
                        pass

# workbook.save('certificados-finalizados/Trena a laser.xlsx')

def capture_output():
    '''
    Function to save the output of the script to a file.

    The file is saved at the 'output' folder.
    '''
    error_list_table = []

    output = StringIO()
    sys.stdout = output

    col1, col2, col3 = st.columns([3, 1, 3])

    with col2:
        image_placeholder = st.empty()
        gif_path = 'images/loading.gif'
        image_placeholder.image(gif_path, use_column_width=True)

    try:
        main()
    except Exception as e:
        (f'{e}')
    try:
        verify_pattern_origin()
    except:
        pass
    try:
        verify_pattern_alignment()
    except:
        pass
    try:
        verify_pattern_font()
    except:
        pass
    try:
        verify_procedure_text_font()
    except:
        pass
    try:
        verify_titles()
    except:
        pass
    try:
        verify_observations_text()
    except:
        pass
    try:
        verify_executer_font()
    except:
        pass
    try:
        verify_table_observation_font()
    except:
        pass
    try:
        verify_header()
    except:
        pass
    try:
        verify_intern_procedure_code()
    except:
        pass
    try:
        verify_Veff()
    except:
        pass
    finally:
        sys.stdout = sys.__stdout__
        output.seek(0)
        image_placeholder.empty()
        return output.getvalue()

output_text = capture_output()  
if error_list:
        col1, col2, col3= st.columns([1, 1.2, 1])
        with col2:
            bullet_list_items = "".join([f"<li>{error}</li>" for error in error_list])
            bullet_list = f"<h2 style='text-align: center; margin-top: 0px'>Erros na formatação do documento:</h2><ul>{bullet_list_items}</ul>"
            st.markdown(bullet_list, unsafe_allow_html=True)
print(output_text)

while output_text is None:
    st.write('Aguarde...')
    break

st.markdown('<div style="display: flex; text-align: center>"' + output_text + '</div>', unsafe_allow_html=True)
