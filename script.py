import pandas as pd 
import openpyxl
import datetime
import nltk
from nltk import tokenize
import requests
from bs4 import BeautifulSoup
from lxml import etree
from lxml import html
import re
import ast
from openpyxl.styles import PatternFill
import numpy as np
import sys

workbook = openpyxl.load_workbook('certificados.xlsx/Trena a laser.xlsx', data_only=True)
sheet = workbook.active

# Get the max row count
max_row = sheet.max_row

# Get the max column count
max_column = sheet.max_column

#Função para pegar a última linha com conteúdo
def get_last_row(sheet):
    for i in range(max_row, 0, -1):
        row_values = [cell.value for cell in sheet[i]]

        if any(row_values):
            return i
        
    return None

last_row = get_last_row(sheet)

#Função para pegar a última coluna com conteúdo
def get_last_column(sheet):	
    for i in range(max_column, 0, -1):
        column_values = [cell.value for cell in sheet[i]]

        if any(column_values):
            return i
        
    return None

def scrapper():
    #Usando o URL apresentado no documento de calibração
    url = 'http://www.inmetro.gov.br/laboratorios/rbc/detalhe_laboratorio.asp?num_certificado=34&situacao=AT&area=DIMENSIONAL'

    response = requests.get(url)

    html_content = response.content

    #Usando o BeautifulSoup para fazer o parse do HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    #Aqui, juntamos o BeautifulSoup com o lxml para fazer o parse do HTML, já que o lxml aceita XPATH
    html_tree = html.fromstring(str(soup))

    #Aqui, usamos o XPATH para pegar a tabela que queremos
    table_rows = html_tree.xpath('//table[4]/tr')

    #Aqui, criamos uma lista vazia para armazenar os dados da tabela
    rows_data = []

    #Aqui, iteramos sobre as linhas da tabela e pegamos o texto de cada célula
    for row in table_rows:
        cells = row.xpath('.//td|.//th')
        row_data = [cell.text_content().strip() for cell in cells]
        rows_data.append(row_data)

    #Aqui, criamos um DataFrame do Pandas com os dados da tabela
    df = pd.DataFrame(rows_data, columns = None, index = None) 

    df = df.dropna(axis=1, how='all') #Aqui, removemos as colunas que só tem valores nulos

    df = df.dropna(axis=0, how='all') #Aqui, removemos as linhas que só tem valores nulos

    # df = df[~df[1].astype(str).str.startswith('Método')]
    # #print(df)

    df = df.replace('Medição de', 'Medir', regex=True)
    df = df.replace('Medição por', 'Medir por', regex=True)
    df = df.replace('para Medir', 'de Medir', regex=True)
    df.columns = ['Descrição do serviço', 'Parâmetro, Faixa e Método', 'Capacidade de Medição e Calibração (CMC)']

    for i, row in df.iterrows():
        if pd.isna(df.at[i, 'Descrição do serviço']) or row['Descrição do serviço'] == '':
            df.at[i, 'Descrição do serviço'] = df.at[i-1, 'Descrição do serviço']

    # df = df.to_excel('web.xlsx', index=False)

    return df

def convert_to_float(value):
    try:
        return float(value.split()[0].replace(',', '.'))
    except:
        return value

def process_string(string):
    result = re.findall(r'\d+', string)

    if result:
        if len(result) != 1:
            return float(result[0]), float(result[1])
        else:
            return 0.0, float(result[0])
    else:
        return None

last_column = get_last_column(sheet)

#print('--------------------------------------------------')

centro_found = False
padroes_found = False
capa_data = []

for i in range (1, last_row + 1):
    row_values = []
    for j in range (1, last_column + 1):
        cell_obj = sheet.cell(row = i, column = j)

        if str(cell_obj.value).startswith('CENTRO'):
            centro_found = True
            continue
        if str(cell_obj.value).startswith('Padrões utilizados'):
            padroes_found = True
            continue

        if centro_found and not padroes_found:
            row_values.append(cell_obj.value)
            if len(row_values) == 9:
                # #print(f'valor da linha {i}: {row_values}')
                capa_data.append(row_values)

    if padroes_found:
        break

def create_df_capa():
    df_capa = pd.DataFrame(capa_data)
    df_capa = df_capa.dropna(axis=1, how='all')
    df_capa = df_capa.dropna(axis=0, how='all')
    df_capa = df_capa[~df_capa.apply(lambda row: 'Ocultar' in row.values, axis=1)]
    df_capa = df_capa.drop_duplicates().reset_index(drop=True)

    return df_capa

df_capa = create_df_capa()

#print("Informações da capa:")
#print(df_capa)


#print('--------------------------------------------------')

#Aqui, vamos tentar extrair os as máquinas utilizadas na medição

padroes_found = False
procedimento_found = False
padroes_data = []

for i in range(1, last_row + 1):
    row_values = []
    for j in range(1, last_column + 1):
        cell_obj = sheet.cell(row = i, column = j)

        if str(cell_obj.value) == 'Padrões utilizados':
            padroes_found = True
            continue

        if str(cell_obj.value).startswith('Procedimento'):
            procedimento_found = True
            continue

        if padroes_found and not procedimento_found:
            row_values.append(cell_obj.value)
            if len(row_values) == 9:
                padroes_data.append(row_values)

    if procedimento_found:
        break

df_padroes = pd.DataFrame(padroes_data)
df_padroes = df_padroes.dropna(axis=1, how='all')

descricao_column = df_padroes[df_padroes.eq('Descrição').any(axis = 1)].stack().index[1][1]

start_row = df_padroes[df_padroes[descricao_column] == 'Descrição'].index[0]
value_below_descricao = df_padroes.loc[start_row + 1][descricao_column]

result_values = [value_below_descricao]
current_row = start_row + 1

while current_row < len(df_padroes) and df_padroes.loc[current_row, descricao_column] is not None:
    result_values.append(df_padroes.loc[current_row, descricao_column])
    current_row += 1

machines_df = pd.DataFrame(result_values, columns=[descricao_column])

machines_df = machines_df.dropna(how='all', inplace=False)
machines_df = machines_df.drop_duplicates().reset_index(drop=True)

machines_df = machines_df[machines_df[descricao_column] != '#N/A']

machines_df.columns = ['Descrição do serviço']

#print('Máquinas utilizadas:')
#print(machines_df)

#print('--------------------------------------------------')

#Nesse bloco de código, fazemos um loop para pegar somente os valores tabelados de medição do certificado

resultados_found = False
observacoes_found = False
data = []

for i in range(1, last_row + 1):
    row_values = []

    for j in range(1, last_column + 1):
        cell_obj = sheet.cell(row = i, column = j)

        if str(cell_obj.value) == 'Resultados':
            resultados_found = True
            continue
    
        if str(cell_obj.value) == 'Observações':
            observacoes_found = True
            continue

        if resultados_found and not observacoes_found:
            row_values.append(cell_obj.value)
            if len(row_values) == 9:
                data.append(row_values)

    if observacoes_found:
        break

df_dados = pd.DataFrame(data)
df_dados = df_dados.dropna(axis=1, how='all')
df_dados = df_dados.dropna(axis=0, how='all')
df_dados = df_dados[~df_dados.apply(lambda row: 'Ocultar' in row.values, axis=1)]
df_dados = df_dados.drop_duplicates().reset_index(drop=True)

#Vamos tentar separar as diferentes tabelas que temos dentro do df_dados

new_table_indices = df_dados[df_dados.apply(lambda row: any(cell and str(cell).startswith('Valor') for cell in row.values), axis=1)].index

# Extract tables
tables = []

for i in range(len(new_table_indices)):
    start_idx = new_table_indices[i]
    end_idx = new_table_indices[i+1] if i+1 < len(new_table_indices) else len(df_dados)
    table = df_dados.iloc[start_idx:end_idx, :].reset_index(drop=True)
    tables.append(table)

#print('Dados da medição:')
#print(df_dados)
#print(tables)

#O índice dentro do [] indica qual tabela específica vamos printar (em ordem de cima para baixo no documento)
#Ou seja, tables separa as tabelas que temos dentro do df_dados

#print('--------------------------------------------------')

df_web = scrapper()
#print('df_web:')
#print(df_web)

#print('--------------------------------------------------')

# Quando damos um merge() nos dataframes, conseguimos um resultado semelhante ao de um PROCV
# Ou seja, aqui estamos fazendo um PROCV entre site e máquina, logo temos a regra da máquina utilizada para cada serviço

df_merge = pd.merge(machines_df, df_web, on='Descrição do serviço', how='left')
df_merge = df_merge.drop_duplicates().reset_index(drop=True)

#print('PROCV entre df_web e machines_df:')
#print(df_merge)

#print('--------------------------------------------------')

df_capa_merge = df_capa.copy()
df_web_merge = df_web.copy()

df_capa_merge[df_capa_merge.columns[1]] = df_capa_merge[df_capa_merge.columns[1]].str.lower()
df_web_merge[df_web_merge.columns[0]] = df_web_merge[df_web_merge.columns[0]].str.lower()

df_service = pd.merge(df_capa_merge, df_web_merge, left_on=df_capa_merge.iloc[:, 1], right_on=df_web_merge.iloc[:, 0], how='inner')
df_service = df_service.drop_duplicates().reset_index(drop=True)



first_column = tables[0].iloc[:, 0]
    
for i, table in enumerate(tables):
    tables[i].iloc[:, 0] = table.iloc[:, 0].apply(convert_to_float)

first_column.columns = ['Resultados']

#print(first_column)

last_value = first_column.iloc[-1]

#Temos o maior valor (sempre o mais inferior da tabela), agora vamos usar ele para saber qual regra usar

# Vamos tentar resolver uma questão sobre trabalhos feitos em campo ou em laboratório
# Imaginamos que, para trabalhos feitos em campo, existe alguma célula com um texto específico


df_web_split = df_web.copy()

indices = df_web_split[df_web_split['Descrição do serviço'] == 'INSTRUMENTOS E GABARITOS DE MEDIÇÃO DE ÂNGULO'].index[1]

df_web_lab = df_web_split.iloc[:indices - 2]
df_web_field = df_web_split.iloc[indices - 2:]

#Vamos descobrir se o certificado é de medição em campo ou em laboratório

if 'LOCAL DA CALIBRAÇÃO' in df_capa.iloc[:, 0].astype(str).values:
    working_df = df_web_field
else:
    working_df = df_web_lab


#Assim, working_df vai sempre armazenar o dataframe correto para utilizarmos no merge()


# Vamos partir para o merge() e assim ter as informações de erro

working_df[working_df.columns[0]] = working_df[working_df.columns[0]].str.lower()

df_merge_service = pd.merge(df_capa_merge, working_df, left_on=df_capa_merge.iloc[:, 1], right_on=working_df.iloc[:, 0], how='inner')

df_merge_service = df_merge_service.drop_duplicates().reset_index(drop=True)
df_merge_service.iloc[:, -1] = df_merge_service.iloc[:, -1].str.replace('*', '')
df_merge_service = df_merge_service.dropna(axis=0, how='all')
df_merge_service = df_merge_service.dropna(axis=1, how='all')

#Vamos para a parte difícil: extrair o intervalo numérico a partir da string


df_merge_service['Intervalo'] = df_merge_service.iloc[:, 4].apply(process_string)
df_merge_service = df_merge_service.dropna(axis=0, how='any')
df_merge_service = df_merge_service.dropna(axis=1, how='any')  

#Agora que temos todas as peças, vamos ao que importa: os cálculos
#Queremos pegar cada valor da first_column e ver em qual intervalo ele se encaixa
#Depois, vamos pegar o valor de erro e incerteza correspondente

def process_cmc_information(cmc_value):
    cmc_value = str(cmc_value).strip()
    
    # Check if it's a distance (type 1)
    distance_match = re.match(r'([\d.,]+)\s*([µm]+)', cmc_value)
    if distance_match:
        #print('Caso 1 utilizado')
        value = float(distance_match.group(1).replace(',', '.'))
        unit = distance_match.group(2)
        return value, unit

    # Check if it's an equation (type 2)
    equation_match = re.match(r'\[([\s\S]+)\]', cmc_value)
    if equation_match:
        #print('Caso 2 utilizado')
        return equation_match.group(1)

    # Check if it's an angle (type 3)
    angle_match = re.match(r'\s*(\d+)\s*\'\'\s*', cmc_value)
    if angle_match:
        #print('Caso 3 utilizado')
        return float(angle_match.group(1))

    # Check if it's a percentage (type 4)
    percentage_match = re.match(r'([\d.,]+)%', cmc_value)
    if percentage_match:
        #print('Caso 4 utilizado')
        return float(percentage_match.group(1).replace(',', '.')) / 100

    # Default case: return the original value
    return cmc_value


df_merge_service['Capacidade de Medição e Calibração (CMC)'] = df_merge_service['Capacidade de Medição e Calibração (CMC)'].apply(process_cmc_information)

def get_error_and_uncertainty(valor, intervalos):
    for intervalo in intervalos:
        if intervalo[0] <= valor <= intervalo[1]:
            return intervalo[0]
    return None

single_row = df_merge_service.iloc[0]

#Aqui, vamos pegar os valores de erro e incerteza e colocar em uma coluna separada

# Iterate through each table in the 'tables' list
for i, row in df_merge_service.iterrows():
    # Iterate through each table in the 'tables' list
    for j, table in enumerate(tables):
        # Get the numerical range for the current row and table
        intervalo = tuple(row['Intervalo'])
        
        # Create a new column with True if the value is within the range, False otherwise
        table[f'Within_Range_{i + 1}'] = pd.to_numeric(table.iloc[:, 0], errors='coerce').between(*intervalo)

#Agora, vamos coorelacionar df_merge_service com as tabelas

for i, table in enumerate(tables):
    # Create a new column to store the selected values
    table['Selected_Value'] = ""

    # Iterate through the rows of the table
    for index, row in table.iterrows():
        selected_value = ""

        # Iterate through the columns and find the first True value in boolean columns
        for col in table.columns:
            if isinstance(col, str) and col.startswith('Within_Range') and row[col]:
                # Extract the range index from the column name
                range_index = int(col.split('_')[-1]) - 1

                # Check if the range_index is within the valid range of rows in df
                if 0 <= range_index < len(df_merge_service):
                    # Get the value from the 6th column of df based on the range_index
                    selected_value = df_merge_service.iloc[range_index, -2]  # Assuming the 6th column index is 5
                    break  # Exit the loop once a match is found

        # Update the 'Selected_Value' column with the corresponding value from df
        table.at[index, 'Selected_Value'] = selected_value

#Aqui, vamos pegar os valores de erro e incerteza e colocar em uma coluna separada

for i, table in enumerate(tables):
    u_column_index = table.columns[table.iloc[0].astype(str).str.startswith('U')].tolist()
    if u_column_index:
        u_column_index = u_column_index[0]
        #print(u_column_index)
        break
    
has_meters = '[m]' in table.iloc[:, u_column_index].values
has_mm = '[mm]' in table.iloc[:, u_column_index].values
has_µm = '[µm]' in table.iloc[:, u_column_index].values


for i in range(len(tables)):
    table = tables[i]

    if 'Selected_Value' in table.columns:
        table[['CMC_Value', 'CMC_Unit']] = table['Selected_Value'].apply(pd.Series)

        table = table.drop('Selected_Value', axis=1)

        tables[i] = table
    else:
        pass

#Agora que temos os valores de CMC e as unidades separadas, podemos trabalhar com as colunas

def convert_to_meters(row):
    value, unit = row['CMC_Value'], row['CMC_Unit']
    if has_meters and unit == 'mm':
        return value / 1000  
    elif has_meters and unit == 'µm':
        return value / 1000000  
    else:
        return value

def convert_to_mm(row):
    value, unit = row['CMC_Value'], row['CMC_Unit']
    if has_mm and unit == 'm':
        return value * 1000  
    elif has_mm and unit == 'µm':
        return value / 1000  
    else:
        return value

def convert_to_µm(row):
    value, unit = row['CMC_Value'], row['CMC_Unit']
    if has_µm and unit == 'm':
        return value * 1000000  
    elif has_µm and unit == 'mm':
        return value / 1000  
    else:
        return value
    
for i in range(len(tables)):
    table = tables[i]
    if 'CMC_Value' in table.columns and 'CMC_Unit' in table.columns and has_meters:
        table['CMC_Value'] = table.apply(convert_to_meters, axis=1)
        tables[i] = table
    elif 'CMC_Value' in table.columns and 'CMC_Unit' in table.columns and has_mm:
        table['CMC_Value'] = table.apply(convert_to_mm, axis=1)
        tables[i] = table
    elif 'CMC_Value' in table.columns and 'CMC_Unit' in table.columns and has_µm:
        table['CMC_Value'] = table.apply(convert_to_µm, axis=1)
        tables[i] = table
    else:
        pass

for i, table in enumerate(tables):
    tables[i].iloc[:, u_column_index] = table.iloc[:, u_column_index].apply(convert_to_float)


new_column_name = 'U'
for i in range(len(tables)):
    tables[i].columns.values[u_column_index] = new_column_name

for i in range(len(tables)):
    table = tables[i]
    table['CMC_Value'] = table['CMC_Value'].replace('', None)
    table['CMC_Verification'] = None
    table['Range_Verification'] = None
    table['Correction_Verification'] = None

search_condition = (df_capa[0] == 'RESOLUÇÃO')
result = df_capa.loc[search_condition]

resolucao_value = result.iloc[0, 1]

resolucao_value = resolucao_value.split()[0]
resolucao_value = resolucao_value.replace(',', '.')
resolucao_value = float(resolucao_value) * 3
resolucao_negative_value = -resolucao_value

for i, table in enumerate(tables):
    for index, row in table.iterrows():
        correction_column_value = row.iloc[2]

        # Check if the value is not None and not NaN
        if pd.notna(correction_column_value) and correction_column_value != 'Correção':
            
            if type(correction_column_value) == str:
                correction_column_value = pd.to_numeric(
                    correction_column_value.replace(',', '.'),
                    errors='coerce'
                    )
            else:
                pass
            
            if pd.notna(correction_column_value):

                if correction_column_value < resolucao_negative_value or correction_column_value > resolucao_value:
                    table.at[index, 'Correction_Verification'] = True
        
            if pd.notna(correction_column_value):  # Check again after conversion
                
                if correction_column_value < resolucao_negative_value or correction_column_value > resolucao_value:

                    table.at[index, 'Correction_Verification'] = True
        else:
            pass

error_ocurred = False

for i, table in enumerate(tables):
    for index, row in table.iterrows():
        u_column_value = row['U']
        cmc_value = row['CMC_Value']

        if pd.notna(cmc_value) and pd.notna(u_column_value):
            if u_column_value >= cmc_value:
                table.at[index, 'CMC_Verification'] = False
            elif u_column_value  < cmc_value:
                table.at[index, 'CMC_Verification'] = True
        elif pd.isna(cmc_value) and type(u_column_value) != str and pd.notna(u_column_value):
            table.at[index, 'Range_Verification'] = True


def find_excel_row_by_value(sheet, target_value):
    max_row = sheet.max_row

    # Iterate through rows in the Excel sheet
    for i in range(1, max_row + 1):
        cell_obj = sheet.cell(row=i, column=1)  

        # Check if the cell contains the target value
        if cell_obj.value == target_value:
            print(f'Match found at row {i}')
            return i  # Return the row number if found

    return None  # Return None if not found


num_rows_painted = 0
# Iterate through each DataFrame in the list
for i, table in enumerate(tables):
    # print(table)

    for index, row in table.iterrows():
        # Assume the first column in the table corresponds to the second column in Excel

        if row['Range_Verification'] == True or row['CMC_Verification'] == True:
            target_value = row[0]
            target_value_str = str(target_value).replace('.', ',')
        # #print(f'Looking for target value: {target_value_str}')
        # Find the row in the Excel file that corresponds to the target value
            excel_row = find_excel_row_by_value(sheet, target_value_str)

            if excel_row is not None and (row['Range_Verification'] or row['CMC_Verification'] or row['Correction_Verification']) == True:
            # Iterate through columns in the DataFrame
                for col_num, value in enumerate(row):
                # Assuming you want to color cells starting from the second column
                    cell_to_paint = sheet.cell(row=excel_row, column=col_num + 1)

                # Set the fill color for the cell
                    fill_color = 'EA4335'  # Specify the color in RGB format (here, red)
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell_to_paint.fill = fill
        
            num_rows_painted += 1
            
        else:
            pass

print(tables)

print('--------------------------------------------------')

def verify_origin():
    standards_df = df_padroes.copy()
    standards_df = standards_df.dropna(axis=1, how='all').reset_index(drop=True)
    standards_df = standards_df.dropna(axis=0, how='all').reset_index(drop=True)

    origin_column = standards_df.iloc[:, -2]
    print(origin_column)

    if origin_column[1] == 'LMD':
        print('Erro: O certificado nao e do padrao CERTI')
    else:
        print('O certificado e do padrao CERTI')

def verify_pattern_alignment():
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Padrões utilizados':
                start_row = cell.row + 1  # Start from the next row
            elif cell.value == 'Procedimento de calibração':
                end_row = cell.row - 1  # End at the previous row
                break

    # Check alignment and content for each cell in the range
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                # Get alignment properties
                alignment = cell.alignment
                horizontal_alignment = alignment.horizontal
                vertical_alignment = alignment.vertical
                # print(f"Cell {cell.coordinate}: Horizontal: {horizontal_alignment}, Vertical: {vertical_alignment}, Content: {cell.value}")
                if horizontal_alignment == vertical_alignment:
                    pass
                else:
                    print(f'Erro de alinhamento: {cell.coordinate}')

def verify_pattern_font():
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Padrões utilizados':
                start_row = cell.row + 2  # Start from the next row
            elif cell.value == 'Procedimento de calibração':
                end_row = cell.row - 1  # End at the previous row
                break

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                # Get alignment properties
                font_name = cell.font.name
                font_size = cell.font.sz
                if font_name == 'Nunito Sans':
                    if cell.row == start_row and font_size == 10:
                        pass
                    elif cell.row == start_row and font_size == 9:
                        print(f'Fonte de cabecalho incorreta {cell.coordinate} -> {font_size}, tamanho de fonte esperado: 10')
                    elif cell.row != start_row and font_size == 9:
                        pass
                    else:
                        print(f'Erro de fonte: {cell.coordinate}, fonte atual: {font_name} - fonte correta: Nunito Sans, tamanho atual: {font_size} - tamanho correto: 9')
                else:
                    print(f'Erro de fonte: {cell.coordinate}, fonte atual: {font_name} - fonte correta: Nunito Sans, tamanho atual: {font_size} - tamanho correto: 9')

def verify_text_font():
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Procedimento de calibração':
                start_row = cell.row + 1  # Start from the next row
            elif cell.value == 'Resultados':
                end_row = cell.row - 1  # End at the previous row
                break

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                font_name = cell.font.name
                font_size = cell.font.sz
                if font_name == 'Nunito Sans' and font_size == 10:
                    pass
                else:
                    print(f'Erro de fonte: {cell.coordinate}, fonte atual: {font_name} - fonte correta: Nunito Sans, tamanho atual: {font_size} - tamanho correto: 10')

def verify_titles():
    titles = []
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Padrões utilizados':
                titles.append(cell.coordinate)
            elif cell.value == 'Procedimento de calibração':
                titles.append(cell.coordinate)
            elif cell.value == 'Resultados':
                titles.append(cell.coordinate)
            elif cell.value == 'Observações':
                titles.append(cell.coordinate)
            else:
                pass
    for item in titles:
        font_name = sheet[item].font.name
        font_size = sheet[item].font.sz
        font_bold = sheet[item].font.b
        if font_name == 'Nunito Sans' and font_size == 11 and font_bold == True:
            pass
        elif font_name != 'Nunito Sans':
            print(f'Erro de fonte: {item}, fonte atual: {font_name} - fonte correta: Nunito Sans')
        elif font_size != 11:
            print(f'Erro de tamanho de fonte: {item}, tamanho atual: {font_size} - tamanho correto: 11')
        elif font_bold != True:
            print(f'Erro de negrito: {item}, negrito atual: {font_bold} - negrito correto: True')
        else:
            pass

def verify_observations():
    start_row = None
    end_row = get_last_row(sheet)

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Observações':
                start_row = cell.row + 1
                pass
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                # Get alignment properties
                font_name = cell.font.name
                font_size = cell.font.sz
                if font_name == 'Nunito Sans' and font_size == 9:
                    pass
                else:
                    print(f'Erro nas observacoes: {cell.coordinate}, fonte atual: {font_name} - fonte correta: Nunito Sans, tamanho atual: {font_size} - tamanho correto: 9.0')


def verify_executer():
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value and str(cell.value).startswith('Executor'):
                start_row = cell.row
            elif cell.value == 'Padrões utilizados':
                end_row = cell.row - 1
                break
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not 
            if cell.value is not None:
                if not str(cell.value).startswith('Executor'):
                    font_name = cell.font.name
                    font_size = cell.font.sz
                    font_bold = cell.font.b
                    if font_name == 'Nunito Sans' and font_size == 10 and font_bold == True:
                        pass
                    elif font_name != 'Nunito Sans':
                        print(f'Erro de fonte: {cell.coordinate}, fonte atual: {font_name} - fonte correta: Nunito Sans')
                    elif font_size != 10:
                        print(f'Erro de tamanho de fonte: {cell.coordinate}, tamanho atual: {font_size} - tamanho correto: 10.0')
                    elif font_bold != True:
                        print(f'Erro de negrito: {cell.coordinate}, negrito atual: {font_bold} - negrito correto: True')
                    else:
                        pass

def verify_table_observation():
    start_row = None
    end_row = None

    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value == 'Resultados':
                start_row = cell.row + 1
            elif cell.value == 'Observações':
                end_row = cell.row - 1
                break
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                if str(cell.value).startswith('Obs'):
                    font_name = cell.font.name
                    font_size = cell.font.sz
                    if font_name == 'Nunito Sans' and font_size == 8:
                        pass
                    elif font_name != 'Nunito Sans':
                        print(f'Erro de fonte: {cell.coordinate}, fonte atual: {font_name} - fonte correta: Nunito Sans')
                    elif font_size != 8:
                        print(f'Erro de tamanho de fonte: {cell.coordinate}, tamanho atual: {font_size} - tamanho correto: 8.0')
                    else:
                        pass

def verify_header():
    start_row_height = sheet.row_dimensions[1].height
    second_row_height = sheet.row_dimensions[2].height
    third_row_height = sheet.row_dimensions[3].height
    fourth_row_height = sheet.row_dimensions[4].height
    fifth_row_height = sheet.row_dimensions[5].height

    start_row_name = sheet.cell(row=1, column=1).font.name
    start_row_size = sheet.cell(row=1, column=1).font.sz
    start_row_bold = sheet.cell(row=1, column=1).font.b

    second_row_name = sheet.cell(row=1 + 1, column=1).font.name
    second_row_size = sheet.cell(row=1 + 1, column=1).font.sz
    second_row_bold = sheet.cell(row=1 + 1, column=1).font.b

    third_row_name = sheet.cell(row=1 + 2, column=1).font.name
    third_row_size = sheet.cell(row=1 + 2, column=1).font.sz
    third_row_bold = sheet.cell(row=1 + 2, column=1).font.b

    fourth_row_name = sheet.cell(row=1 + 3, column=1).font.name
    fourth_row_size = sheet.cell(row=1 + 3, column=1).font.sz
    fourth_row_bold = sheet.cell(row=1 + 3, column=1).font.b

    fifth_row_name = sheet.cell(row=1 + 4, column=1).font.name
    fifth_row_size = sheet.cell(row=1 + 4, column=1).font.sz
    fifht_row_bold = sheet.cell(row=1 + 4, column=1).font.b

    if start_row_name == 'Nunito Sans' and start_row_size == 14 and start_row_height == 24.75 and start_row_bold != True:
        pass
    elif start_row_name != 'Nunito Sans':
        print(f'Erro de fonte: {sheet.cell(row=start_row, column=1).coordinate}, fonte atual: {start_row_name} - fonte correta: Nunito Sans')
    elif start_row_size != 14:
        print(f'Erro de tamanho de fonte: {sheet.cell(row=start_row, column=1).coordinate}, tamanho atual: {start_row_size} - tamanho correto: 14.0')
    elif start_row_height != 24.75:
        print(f'Erro de altura de linha: {sheet.cell(row=start_row, column=1).coordinate}, altura atual: {start_row_height} - altura correta: 24.75')
    elif start_row_bold == True:
        print(f'Erro de negrito: {sheet.cell(row=start_row, column=1).coordinate}, negrito atual: {start_row_bold} - negrito correto: False')
    else:
        pass

    if second_row_name == 'Nunito Sans' and second_row_size == 14 and second_row_height == 17.25 and second_row_bold != True:
        pass
    elif second_row_name != 'Nunito Sans':
        print(f'Erro de fonte: {sheet.cell(row=1 + 1, column=1).coordinate}, fonte atual: {second_row_name} - fonte correta: Nunito Sans')
    elif second_row_size != 14:
        print(f'Erro de tamanho de fonte: {sheet.cell(row=1 + 1, column=1).coordinate}, tamanho atual: {second_row_size} - tamanho correto: 11.0')
    elif second_row_height != 17.25:
        print(f'Erro de altura de linha: {sheet.cell(row=1 + 1, column=1).coordinate}, altura atual: {second_row_height} - altura correta: 18.0')
    elif second_row_bold == True:
        print(f'Erro de negrito: {sheet.cell(row=1 + 1, column=1).coordinate}, negrito atual: {second_row_bold} - negrito correto: False')
    else:
        pass

    if third_row_name == 'Nunito Sans' and third_row_size == 9 and third_row_height == 15 and third_row_bold != True:
        pass
    elif third_row_name != 'Nunito Sans':
        print(f'Erro de fonte: {sheet.cell(row=1 + 2, column=1).coordinate}, fonte atual: {third_row_name} - fonte correta: Nunito Sans')
    elif third_row_size != 9:
        print(f'Erro de tamanho de fonte: {sheet.cell(row=1 + 2, column=1).coordinate}, tamanho atual: {third_row_size} - tamanho correto: 11.0')
    elif third_row_height != 15:
        print(f'Erro de altura de linha: {sheet.cell(row=1 + 2, column=1).coordinate}, altura atual: {third_row_height} - altura correta: 18.0')
    elif third_row_bold == True:
        print(f'Erro de negrito: {sheet.cell(row=1 + 2, column=1).coordinate}, negrito atual: {third_row_bold} - negrito correto: False')
    else:
        pass

    if fourth_row_name == 'Nunito Sans' and fourth_row_size == 22 and fourth_row_height == 39 and fourth_row_bold == True:
        pass
    elif fourth_row_name != 'Nunito Sans':
        print(f'Erro de fonte: {sheet.cell(row=1 + 3, column=1).coordinate}, fonte atual: {fourth_row_name} - fonte correta: Nunito Sans')
    elif fourth_row_size != 22:
        print(f'Erro de tamanho de fonte: {sheet.cell(row=1 + 3, column=1).coordinate}, tamanho atual: {fourth_row_size} - tamanho correto: 22.0')
    elif fourth_row_height != 39:
        print(f'Erro de altura de linha: {sheet.cell(row=1 + 3, column=1).coordinate}, altura atual: {fourth_row_height} - altura correta: 39.0')
    elif fourth_row_bold != True:
        print(f'Erro de negrito: {sheet.cell(row=1 + 3, column=1).coordinate}, negrito atual: {fourth_row_bold} - negrito correto: True')
    else:
        pass

    if fifth_row_name == 'Nunito Sans' and fifth_row_size == 22 and fifth_row_height == 35.25 and fifht_row_bold == True:
        pass
    elif fifth_row_name != 'Nunito Sans':
        print(f'Erro de fonte: {sheet.cell(row=1 + 4, column=1).coordinate}, fonte atual: {fifth_row_name} - fonte correta: Nunito Sans')
    elif fifth_row_size != 22:
        print(f'Erro de tamanho de fonte: {sheet.cell(row=1 + 4, column=1).coordinate}, tamanho atual: {fifth_row_size} - tamanho correto: 22.0')
    elif fifth_row_height != 35.25:
        print(f'Erro de altura de linha: {sheet.cell(row=1 + 4, column=1).coordinate}, altura atual: {fifth_row_height} - altura correta: 35.25')
    elif fifht_row_bold != True:
        print(f'Erro de negrito: {sheet.cell(row=1 + 4, column=1).coordinate}, negrito atual: {fifht_row_bold} - negrito correto: True')
    else:
        pass

# workbook.save('certificados-finalizados/Trena a laser.xlsx')

def save_output_to_file(file_path):
    original_stdout = sys.stdout
    try:
        with open(file_path, 'w') as f:
            sys.stdout = f
            verify_origin()
            verify_pattern_alignment()
            verify_pattern_font()
            verify_text_font()
            verify_titles()
            verify_observations()
            verify_executer()
            verify_table_observation()
            verify_header()
    finally:
        sys.stdout = original_stdout

output_file_path = 'output/output.txt'

