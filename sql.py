import pyodbc
import pandas as pd
import openpyxl


conn_str = ("Driver={SQL Server};"
            "Server=flndtcvmsql01\certi;"
            "Database=CALI;"
            "Trusted_Connection=yes;")

conn = pyodbc.connect(conn_str)
# cursor = conn.cursor()
user_sql = "SELECT DISTINCT USUARIO.NMUSUARIO, USUARIOLAB.TITULO, USUARIO.COLABORADOR FROM CALIBRACAO JOIN USUARIO ON CALIBRACAO.CDUSUARIO = USUARIO.CDUSUARIO JOIN USUARIOLAB ON USUARIO.CDUSUARIO = USUARIOLAB.CDUSUARIO WHERE TITULO != 'Ex-colaborador'"

cali_sql = "SELECT C.CDCALIBRACAO, C.CDFORNECEDOR, C.CDSOLICITANTE,C.CDCONTRATANTE, S.NMFANTASIA AS NMFANTASIASOLICITANTE, F.NMFANTASIA AS NMFANTASIAFORNECEDOR, CT.NMFANTASIA AS NMFANTASIACONTRATANTE FROM CALIBRACAO C JOIN CLIENTE S ON S.CDCLIENTE = C.CDSOLICITANTE JOIN CLIENTE F ON F.CDCLIENTE = C.CDFORNECEDOR JOIN CLIENTE CT ON CT.CDCLIENTE = C.CDCONTRATANTE"

user_data = pd.read_sql(user_sql, conn)
# print(data)

cali_data = pd.read_sql(cali_sql, conn)
print(cali_data)

workbook = openpyxl.load_workbook('certificados.xlsx/Trena a laser.xlsx', data_only=True)
sheet = workbook.active

# Get the max row count
max_row = sheet.max_row

# Get the max column count
max_column = sheet.max_column

#Função para pegar a última linha com conteúdo
def get_last_row(sheet):
    for i in range(max_row, 0, -1):
        row_values = [cell.value for cell in sheet[i]]

        if any(row_values):
            return i

    return None

last_row = get_last_row(sheet)

#Função para pegar a última coluna com conteúdo
def get_last_column(sheet):
    for i in range(max_column, 0, -1):
        column_values = [cell.value for cell in sheet[i]]

        if any(column_values):
            return i

    return None

last_column = get_last_column(sheet)

centro_found = False
padroes_found = False
capa_data = []

for i in range (1, last_row + 1):
    row_values = []
    for j in range (1, last_column + 1):
        cell_obj = sheet.cell(row = i, column = j)

        if str(cell_obj.value).startswith('CENTRO'):
            centro_found = True
            continue
        if str(cell_obj.value).startswith('Padrões utilizados'):
            padroes_found = True
            continue

        if centro_found and not padroes_found:
            row_values.append(cell_obj.value)
            if len(row_values) == 9:
                # print(f'valor da linha {i}: {row_values}')
                capa_data.append(row_values)

    if padroes_found:
        break

def create_df_capa(oi = 1):
    print(oi)
    df_capa = pd.DataFrame(capa_data)
    df_capa = df_capa.dropna(axis=1, how='all')
    df_capa = df_capa.dropna(axis=0, how='all')
    df_capa = df_capa[~df_capa.apply(lambda row: 'Ocultar' in row.values, axis=1)]
    df_capa = df_capa.drop_duplicates().reset_index(drop=True)

    return df_capa

df_capa = create_df_capa()
print(df_capa)

def collaborator_compare():
    merged_df = pd.merge(user_data, df_capa, left_on='NMUSUARIO', right_on=2, how='inner')
    return merged_df

# print(collaborator_compare())

def enterpriser_compare():

### criar função que chama o create_df_capa(), acha a linha do CONTRATANTE, splita a célula
### do contrante + endereço e faz um merge com a cali_data usando só o contrante

### já começa o dia conectando esse vs code ao teu github e fazendo um push

    return df_capa

print(enterpriser_compare())

